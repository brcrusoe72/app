from pathlib import Path

from openpyxl import load_workbook


def test_workbook_contract():
    wb_path = Path(__file__).resolve().parents[1] / "excel" / "Shift_Flight_Deck.xlsm"
    if not wb_path.exists():
        import subprocess
        subprocess.check_call(["python", str(Path(__file__).resolve().parents[1] / "scripts" / "build_or_repair_workbook.py")])
    wb = load_workbook(wb_path, keep_vba=True)

    required_sheets = [
        "Parameters",
        "Schedule_Entry",
        "Hourly_Log",
        "Downtime_Log",
        "Dash_Shift",
        "Dash_Trends",
        "Profiles",
        "Analysis_Report",
        "Rules_Authoring",
    ]
    for s in required_sheets:
        assert s in wb.sheetnames

    assert "tblRules" in wb["Rules_Authoring"].tables
    rules_headers = [wb["Rules_Authoring"].cell(1, c).value for c in range(1, 18)]
    assert "IfLogic" in rules_headers


def test_analyzer_smoke():
    import importlib.util
    mod_path = Path(__file__).resolve().parents[1] / "scripts" / "analyze_workbook.py"
    import sys
    spec = importlib.util.spec_from_file_location("analyze_workbook", mod_path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules["analyze_workbook"] = module
    spec.loader.exec_module(module)
    parse_iflogic = module.parse_iflogic

    parts = parse_iflogic('MISSING_STANDARD(groupby="Line,SKU_Resolved")')
    assert parts[0][0] == "MISSING_STANDARD"
