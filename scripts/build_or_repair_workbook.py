#!/usr/bin/env python3
"""Create or repair the Shift Flight Deck workbook with deterministic structure."""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = REPO_ROOT / "excel" / "Shift_Flight_Deck.xlsm"
RULES_JSON = REPO_ROOT / "data" / "rules.json"

SHEETS = [
    "Parameters",
    "Schedule_Entry",
    "Hourly_Log",
    "Downtime_Log",
    "Dash_Shift",
    "Dash_Trends",
    "Profiles",
    "Analysis_Report",
    "Rules_Authoring",
]

TABLE_DEFS = {
    "tblLines": ("Parameters", ["Line", "TargetRateAttain", "ShiftStartTime", "ShiftEndTime", "Notes"]),
    "tblStandards": ("Parameters", ["Line", "SKU", "ProductName", "Std_CPH"]),
    "tblMachines": ("Parameters", ["Line", "Machine"]),
    "tblOperators": ("Parameters", ["EmpID", "OperatorName", "Role", "TrainedLines"]),
    "tblSchedule": ("Schedule_Entry", ["RowID", "Date", "Shift", "Line", "StartDT", "EndDT", "Order", "SKU", "PlannedCases", "Notes"]),
    "tblHourly": ("Hourly_Log", ["RowID", "Date", "Shift", "Line", "HourEndingDT", "ActualCases", "SKU_Resolved", "Std_CPH", "StdCasesThisHour", "RateAttain_100", "TargetRateAttain", "TargetAttain"]),
    "tblDowntime": ("Downtime_Log", ["RowID", "Date", "Shift", "Line", "StartDT", "EndDT", "Minutes", "Machine", "OperatorEmpID", "Category", "Cause", "ActionTaken", "EscalatedYN", "ResolvedBy", "Notes"]),
    "tblRules": (
        "Rules_Authoring",
        [
            "RuleID", "Enabled", "Severity", "Scope", "Description", "IfLogic", "ThenRecommendation", "ThenEscalation",
            "Thresholds", "WindowHours", "ConsecutiveHours", "AppliesToLine", "AppliesToMachine", "AppliesToSKU", "Version",
            "LastEditedBy", "LastEditedDT",
        ],
    ),
}

DEFAULT_RULES = [
    {
        "RuleID": "R1_UNDERPERFORM_STOPS",
        "Enabled": "TRUE",
        "Severity": "Action",
        "Scope": "Line",
        "Description": "Sustained underperformance with frequent stops",
        "IfLogic": 'CONSEC_BELOW(metric="TargetAttain", threshold=0.70, hours=2, groupby="Line") AND ROLLING_COUNT(table="Downtime", window_hours=2, where="Line={Line}", min=4)',
        "ThenRecommendation": "Run rapid loss review and assign immediate support to top downtime cause.",
        "ThenEscalation": "Notify area lead if persists for 2 more hours.",
        "Thresholds": '{"threshold":0.70,"min_stops":4}',
        "WindowHours": 2,
        "ConsecutiveHours": 2,
        "AppliesToLine": "*",
        "AppliesToMachine": "*",
        "AppliesToSKU": "*",
        "Version": 1,
        "LastEditedBy": "system",
        "LastEditedDT": dt.datetime.now().isoformat(timespec="minutes"),
    },
    {
        "RuleID": "R2_MISSING_STANDARD",
        "Enabled": "TRUE",
        "Severity": "Urgent",
        "Scope": "Line",
        "Description": "Standards missing for active run",
        "IfLogic": 'MISSING_STANDARD(groupby="Line,SKU_Resolved")',
        "ThenRecommendation": "Add standard immediately or switch to approved alternate SKU standard.",
        "ThenEscalation": "Escalate to process engineer and planner.",
        "Thresholds": "{}",
        "WindowHours": 4,
        "ConsecutiveHours": 1,
        "AppliesToLine": "*",
        "AppliesToMachine": "*",
        "AppliesToSKU": "*",
        "Version": 1,
        "LastEditedBy": "system",
        "LastEditedDT": dt.datetime.now().isoformat(timespec="minutes"),
    },
]


def row_id(*parts: object) -> str:
    raw = "|".join(str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def ensure_sheet(wb: Workbook, name: str):
    if name not in wb.sheetnames:
        wb.create_sheet(name)
    return wb[name]


def clear_table(ws, table_name: str):
    if table_name in ws.tables:
        del ws.tables[table_name]


def write_table(ws, table_name: str, columns: list[str], rows: list[list[object]], start_row: int = 1):
    for i, col in enumerate(columns, start=1):
        ws.cell(start_row, i, col).font = Font(bold=True)
    for r, row in enumerate(rows, start=start_row + 1):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    end_row = max(start_row + 1, start_row + len(rows))
    end_col = len(columns)
    ref = f"A{start_row}:{chr(64 + end_col)}{end_row}"
    clear_table(ws, table_name)
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
    ws.add_table(tab)


def ensure_validations(wb: Workbook):
    ws_rules = wb["Rules_Authoring"]
    dv_enabled = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)
    dv_severity = DataValidation(type="list", formula1='"Info,Watch,Action,Urgent"', allow_blank=False)
    dv_scope = DataValidation(type="list", formula1='"Line,Machine,Operator,Shift"', allow_blank=False)
    ws_rules.add_data_validation(dv_enabled)
    ws_rules.add_data_validation(dv_severity)
    ws_rules.add_data_validation(dv_scope)
    dv_enabled.add("B2:B1000")
    dv_severity.add("C2:C1000")
    dv_scope.add("D2:D1000")


def seed_defaults(wb: Workbook):
    ws = wb["Parameters"]
    lines = [[f"Line {i}", 0.85, "06:00", "18:00", ""] for i in range(1, 6)]
    write_table(ws, "tblLines", TABLE_DEFS["tblLines"][1], lines)

    std_rows = []
    for line in [f"Line {i}" for i in range(1, 6)]:
        for n in range(1, 3):
            std_rows.append([line, f"SKU-{n:03d}", f"Product {n}", 100 + n * 10])
    write_table(ws, "tblStandards", TABLE_DEFS["tblStandards"][1], std_rows)

    machines = [[f"Line {i}", f"M{i}-{j}"] for i in range(1, 6) for j in range(1, 3)]
    write_table(ws, "tblMachines", TABLE_DEFS["tblMachines"][1], machines)

    operators = [[f"E{100+i}", f"Operator {i}", "Operator", "Line 1,Line 2"] for i in range(1, 11)]
    write_table(ws, "tblOperators", TABLE_DEFS["tblOperators"][1], operators)

    d = dt.date.today()
    sched_rows = [
        [row_id(d, "A", "Line 1", "SKU-001", "06:00"), d.isoformat(), "A", "Line 1", f"{d} 06:00", f"{d} 10:00", "ORD-1001", "SKU-001", 400, ""],
        [row_id(d, "A", "Line 1", "SKU-002", "10:00"), d.isoformat(), "A", "Line 1", f"{d} 10:00", f"{d} 14:00", "ORD-1002", "SKU-002", 420, ""],
    ]
    write_table(wb["Schedule_Entry"], "tblSchedule", TABLE_DEFS["tblSchedule"][1], sched_rows)

    hourly_rows = []
    for h in [7, 8, 9]:
        rid = row_id(d, "A", "Line 1", f"{h}:00")
        hourly_rows.append([rid, d.isoformat(), "A", "Line 1", f"{d} {h:02d}:00", 85 + h, "SKU-001", 110, 110, f"=(F{len(hourly_rows)+2}/I{len(hourly_rows)+2})", 0.85, f"=(J{len(hourly_rows)+2}/K{len(hourly_rows)+2})"])
    write_table(wb["Hourly_Log"], "tblHourly", TABLE_DEFS["tblHourly"][1], hourly_rows)

    dt_rows = [
        [row_id(d, "Line 1", "M1-1", "07:05"), d.isoformat(), "A", "Line 1", f"{d} 07:05", f"{d} 07:20", 15, "M1-1", "E101", "Mechanical", "Jam", "Cleared", "Y", "Lead", ""],
    ]
    write_table(wb["Downtime_Log"], "tblDowntime", TABLE_DEFS["tblDowntime"][1], dt_rows)

    write_table(wb["Rules_Authoring"], "tblRules", TABLE_DEFS["tblRules"][1], [[r.get(c, "") for c in TABLE_DEFS["tblRules"][1]] for r in DEFAULT_RULES])


def setup_dashboards(wb: Workbook):
    dash = wb["Dash_Shift"]
    dash["A1"] = "Shift Flight Deck"
    dash["A1"].font = Font(size=16, bold=True)
    dash["A3"] = "Data Quality Score"
    dash["B3"] = "=IFERROR(AVERAGE(B6:B8),0)"
    dash["A6"] = "% Hourly with SKU"
    dash["A7"] = "% Standards present"
    dash["A8"] = "% Downtime required fields"
    dash["B6"] = "=IFERROR(COUNTA(Hourly_Log!G:G)/MAX(COUNTA(Hourly_Log!A:A)-1,1),0)"
    dash["B7"] = "=0"
    dash["B8"] = '=IFERROR(COUNTIFS(Downtime_Log!A:A,"<>",Downtime_Log!H:H,"<>",Downtime_Log!J:J,"<>")/MAX(COUNTA(Downtime_Log!A:A)-1,1),0)'
    dash["A10"] = "Changeover Prep Needed"
    dash["B10"] = "=\"Check schedule in next 90 minutes\""

    trends = wb["Dash_Trends"]
    trends["A1"] = "Trend Dashboard"
    trends["A2"] = "Generated by automation scripts"

    report = wb["Analysis_Report"]
    report["A1"] = "Analysis Report"
    report["A2"] = "Run Analyze (Deep) to refresh"

    rules = wb["Rules_Authoring"]
    rules["T1"] = "Rule Linter"
    rules["T2"] = "Issues are written by analyzer"


def try_inject_vba_and_buttons(workbook_path: Path):
    try:
        import win32com.client  # type: ignore
    except Exception:
        return "win32com unavailable; VBA/button injection skipped"

    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    wb = xl.Workbooks.Open(str(workbook_path))
    try:
        module = wb.VBProject.VBComponents.Add(1)
        module.Name = "ShiftDeckMacros"
        code = '''
Public Sub AnalyzeDeep()
    RunPy "scripts\\analyze_workbook.py --workbook """ & ActiveWorkbook.FullName & """ --rules ""data\\rules.json"""
End Sub
Public Sub ArchiveData()
    RunPy "scripts\\archive_history.py --workbook """ & ActiveWorkbook.FullName & """"
End Sub
Public Sub PublishOutputs()
    RunPy "scripts\\publish_reports.py --workbook """ & ActiveWorkbook.FullName & """"
End Sub
Public Sub ExportRules()
    RunPy "scripts\\analyze_workbook.py --workbook """ & ActiveWorkbook.FullName & """ --export-rules"
End Sub
Private Sub RunPy(args As String)
    Dim sh As Object
    Set sh = CreateObject("WScript.Shell")
    sh.Run "cmd /c python " & args, 0, False
End Sub
'''
        module.CodeModule.AddFromString(code)
        wb.Save()
        return "win32com VBA injection completed"
    finally:
        wb.Close(SaveChanges=True)
        xl.Quit()


def export_default_rules_json():
    RULES_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = {"generated_at": dt.datetime.now().isoformat(), "rules": DEFAULT_RULES}
    RULES_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def build_or_repair(path: Path):
    if path.exists():
        wb = load_workbook(path, keep_vba=True)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    for name in SHEETS:
        ensure_sheet(wb, name)

    for table_name, (sheet_name, cols) in TABLE_DEFS.items():
        ws = wb[sheet_name]
        if table_name not in ws.tables:
            write_table(ws, table_name, cols, [])

    if wb["Parameters"].max_row <= 1:
        seed_defaults(wb)

    ensure_validations(wb)
    setup_dashboards(wb)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    export_default_rules_json()
    return try_inject_vba_and_buttons(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(WORKBOOK_PATH))
    args = parser.parse_args()
    msg = build_or_repair(Path(args.workbook))
    print(f"Workbook ready: {args.workbook}")
    print(msg)


if __name__ == "__main__":
    main()
