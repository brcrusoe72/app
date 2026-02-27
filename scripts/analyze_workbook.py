#!/usr/bin/env python3
"""Deep analysis and deterministic rules engine for Shift Flight Deck."""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = REPO_ROOT / "excel" / "Shift_Flight_Deck.xlsm"
DEFAULT_RULES_JSON = REPO_ROOT / "data" / "rules.json"
LOG_DIR = REPO_ROOT / "data" / "logs"

SEVERITY_ORDER = {"Urgent": 4, "Action": 3, "Watch": 2, "Info": 1}
REQ_RULE_COLS = [
    "RuleID", "Enabled", "Severity", "Scope", "Description", "IfLogic", "ThenRecommendation", "ThenEscalation",
    "Thresholds", "WindowHours", "ConsecutiveHours", "AppliesToLine", "AppliesToMachine", "AppliesToSKU", "Version",
    "LastEditedBy", "LastEditedDT",
]


@dataclass
class Trigger:
    rule_id: str
    severity: str
    trigger: str
    evidence: str
    recommendation: str
    scope: str
    affected_entity: str
    timestamp: str
    impact: float


def table_rows(ws, table_name: str) -> list[dict[str, Any]]:
    tab = ws.tables[table_name]
    min_cell, max_cell = tab.ref.split(":")
    min_col = ord(min_cell[0]) - 64
    min_row = int(min_cell[1:])
    max_col = ord(max_cell[0]) - 64
    max_row = int(max_cell[1:])
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    data = []
    for r in range(min_row + 1, max_row + 1):
        row = {headers[i]: ws.cell(r, min_col + i).value for i in range(len(headers))}
        if any(v not in (None, "") for v in row.values()):
            row["_sheet_row"] = r
            data.append(row)
    return data


def parse_call(expr: str) -> tuple[str, dict[str, Any]]:
    m = re.match(r"([A-Z_]+)\((.*)\)", expr.strip())
    if not m:
        raise ValueError(f"Invalid DSL expression: {expr}")
    fn, raw = m.group(1), m.group(2)
    args: dict[str, Any] = {}
    for part in re.split(r",\s*(?=[a-zA-Z_]+=)", raw):
        if not part.strip():
            continue
        k, v = part.split("=", 1)
        v = v.strip().strip('"')
        if re.match(r"^-?\d+\.\d+$", v):
            cast: Any = float(v)
        elif re.match(r"^-?\d+$", v):
            cast = int(v)
        else:
            cast = v
        args[k.strip()] = cast
    return fn, args




def to_float(v: Any) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    if v in (None, ""):
        return 0.0
    txt = str(v).strip()
    if txt.startswith("="):
        return 0.0
    try:
        return float(txt)
    except ValueError:
        return 0.0

def parse_iflogic(iflogic: str) -> list[tuple[str, dict[str, Any]]]:
    chunks = [c.strip() for c in re.split(r"\s+AND\s+", iflogic)]
    return [parse_call(c) for c in chunks if c]


def rolling_count(events: list[dict[str, Any]], window_hours: int, by: list[str]) -> dict[tuple, int]:
    cutoff = dt.datetime.now() - dt.timedelta(hours=window_hours)
    counts: dict[tuple, int] = {}
    for e in events:
        t = parse_dt(e.get("StartDT") or e.get("HourEndingDT"))
        if t and t >= cutoff:
            key = tuple(e.get(k) for k in by)
            counts[key] = counts.get(key, 0) + 1
    return counts


def consecutive_below(metric_series: list[dict[str, Any]], threshold: float, consecutive_hours: int, groupby: list[str], metric: str) -> list[tuple]:
    hits = []
    grouped: dict[tuple, list[float]] = {}
    for row in metric_series:
        key = tuple(row.get(g) for g in groupby)
        grouped.setdefault(key, []).append(to_float(row.get(metric)))
    for key, values in grouped.items():
        streak = 0
        for v in values[-consecutive_hours * 2:]:
            streak = streak + 1 if v < threshold else 0
            if streak >= consecutive_hours:
                hits.append(key)
                break
    return hits


def repeats_same_value(rows: list[dict[str, Any]], field: str, min_repeats: int, window_hours: int, by: list[str]) -> list[tuple]:
    recent = rolling_count(rows, window_hours, by + [field])
    return [k for k, c in recent.items() if c >= min_repeats]


def forecast_shortfall(planned: float, forecast: float, pct_threshold: float) -> bool:
    if planned <= 0:
        return False
    return ((planned - forecast) / planned) >= pct_threshold


def missing_standard(line: str, sku: str, standards: set[tuple[str, str]]) -> bool:
    return (line, sku) not in standards


def schedule_overlap(schedule_rows: list[dict[str, Any]], line: str) -> bool:
    slots = sorted(
        [(parse_dt(r.get("StartDT")), parse_dt(r.get("EndDT"))) for r in schedule_rows if r.get("Line") == line],
        key=lambda x: x[0] or dt.datetime.min,
    )
    for i in range(1, len(slots)):
        if slots[i - 1][1] and slots[i][0] and slots[i][0] < slots[i - 1][1]:
            return True
    return False


def missing_schedule_for_hourly(schedule_rows: list[dict[str, Any]], line: str, hour: dt.datetime) -> bool:
    for r in schedule_rows:
        if r.get("Line") != line:
            continue
        st, en = parse_dt(r.get("StartDT")), parse_dt(r.get("EndDT"))
        if st and en and st <= hour <= en:
            return False
    return True


def outlier_rate(actual: float, std: float, z_or_pct_threshold: float) -> bool:
    if std <= 0:
        return False
    return abs(actual - std) / std >= z_or_pct_threshold


def parse_dt(v: Any):
    if isinstance(v, dt.datetime):
        return v
    if not v:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(str(v), fmt)
        except ValueError:
            pass
    return None


def lint_rules(rules: list[dict[str, Any]]) -> list[str]:
    issues: list[str] = []
    ids = [r.get("RuleID") for r in rules]
    if len(set(ids)) != len(ids):
        issues.append("Duplicate RuleID values detected")
    for i, r in enumerate(rules, start=2):
        missing = [c for c in REQ_RULE_COLS if c not in r or r[c] in (None, "")]
        if missing and str(r.get("Enabled", "")).upper() == "TRUE":
            issues.append(f"Row {i}: missing required fields {missing}")
        if r.get("Severity") not in SEVERITY_ORDER:
            issues.append(f"Row {i}: invalid Severity")
        if r.get("Scope") not in {"Line", "Machine", "Operator", "Shift"}:
            issues.append(f"Row {i}: invalid Scope")
        try:
            parse_iflogic(str(r.get("IfLogic", "")))
        except Exception as exc:
            issues.append(f"Row {i}: DSL parse error {exc}")
    return issues


def sanitize_recommendation(text: str) -> str:
    banned = ["disciplinary", "write-up", "punish", "terminate"]
    if any(b in text.lower() for b in banned):
        return "Provide coaching and process support to remove the operational barrier."
    return text


def evaluate_rules(rules, schedule_rows, hourly_rows, downtime_rows, standards_rows) -> list[Trigger]:
    standards = {(r.get("Line"), r.get("SKU")) for r in standards_rows}
    triggers: list[Trigger] = []

    for rule in rules:
        if str(rule.get("Enabled", "")).upper() != "TRUE":
            continue
        parsed = parse_iflogic(str(rule.get("IfLogic", "")))
        rule_hits = []

        for fn, args in parsed:
            if fn == "CONSEC_BELOW":
                group = str(args.get("groupby", "Line")).split(",")
                hits = consecutive_below(hourly_rows, float(args.get("threshold", 0.7)), int(args.get("hours", 2)), group, str(args.get("metric", "TargetAttain")))
                rule_hits.append(set(hits))
            elif fn == "ROLLING_COUNT":
                group = [g for g in str(args.get("where", "Line={Line}")).replace("={Line}", "").split(",") if g]
                counts = rolling_count(downtime_rows, int(args.get("window_hours", 2)), group or ["Line"])
                hit = {k for k, v in counts.items() if v >= int(args.get("min", 1))}
                rule_hits.append(hit)
            elif fn == "MISSING_STANDARD":
                hit = {(r.get("Line"), r.get("SKU_Resolved")) for r in hourly_rows if missing_standard(r.get("Line"), r.get("SKU_Resolved"), standards)}
                rule_hits.append(hit)
            elif fn == "SCHEDULE_OVERLAP":
                lines = {r.get("Line") for r in schedule_rows}
                hit = {(l,) for l in lines if schedule_overlap(schedule_rows, l)}
                rule_hits.append(hit)
            elif fn == "REPEAT_CAUSE":
                group = str(args.get("groupby", "Line,Machine,Cause")).split(",")
                hit = set(repeats_same_value(downtime_rows, "Cause", int(args.get("min_repeats", 3)), int(args.get("window_hours", 12)), group[:-1]))
                rule_hits.append(hit)
            elif fn == "FORECAST_SHORTFALL":
                by_line = {}
                for r in hourly_rows:
                    by_line.setdefault(r.get("Line"), []).append(to_float(r.get("ActualCases")))
                planned_by_line = {}
                for r in schedule_rows:
                    planned_by_line[r.get("Line")] = planned_by_line.get(r.get("Line"), 0) + to_float(r.get("PlannedCases"))
                hit = set()
                for line, vals in by_line.items():
                    rolling = sum(vals[-3:]) / max(min(3, len(vals)), 1)
                    forecast = sum(vals) + rolling * 2
                    if forecast_shortfall(planned_by_line.get(line, 0), forecast, float(args.get("pct", 0.1))):
                        hit.add((line,))
                rule_hits.append(hit)

        if not rule_hits:
            continue
        inter = set.intersection(*rule_hits) if len(rule_hits) > 1 else set(rule_hits[0])
        for h in inter:
            entity = ",".join(str(x) for x in h if x not in (None, ""))
            now = dt.datetime.now().isoformat(timespec="seconds")
            recommendation = sanitize_recommendation(str(rule.get("ThenRecommendation", "")))
            triggers.append(Trigger(rule.get("RuleID", ""), rule.get("Severity", "Info"), str(rule.get("Description", "")), str(rule.get("IfLogic", "")), recommendation, rule.get("Scope", "Line"), entity or "Unknown", now, float(len(entity))))

    triggers.sort(key=lambda t: (-SEVERITY_ORDER.get(t.severity, 0), -t.impact, t.timestamp), reverse=False)
    return triggers


def select_rules(wb, rules_json: Path):
    ws = wb["Rules_Authoring"]
    rules = table_rows(ws, "tblRules")
    if rules:
        return rules, "workbook"
    if rules_json.exists():
        payload = json.loads(rules_json.read_text(encoding="utf-8"))
        return payload.get("rules", []), "json"
    from scripts.build_or_repair_workbook import DEFAULT_RULES

    return DEFAULT_RULES, "default"


def write_analysis_report(wb, sections: dict[str, list[str]], triggers: list[Trigger], lint_issues: list[str]):
    ws = wb["Analysis_Report"]
    ws.delete_rows(1, ws.max_row)
    row = 1
    ws.cell(row, 1, "Analysis Report")
    row += 2
    for title, lines in sections.items():
        ws.cell(row, 1, title)
        row += 1
        for line in lines:
            ws.cell(row, 1, f"- {line}")
            row += 1
        row += 1

    ws.cell(row, 1, "Rules Engine Coaching Prompts")
    row += 1
    ws.cell(row, 1, "RuleID")
    ws.cell(row, 2, "Severity")
    ws.cell(row, 3, "Trigger")
    ws.cell(row, 4, "Evidence")
    ws.cell(row, 5, "Recommendation")
    ws.cell(row, 6, "Scope")
    ws.cell(row, 7, "AffectedEntity")
    ws.cell(row, 8, "Timestamp")
    row += 1
    for t in triggers:
        ws.cell(row, 1, t.rule_id)
        ws.cell(row, 2, t.severity)
        ws.cell(row, 3, t.trigger)
        ws.cell(row, 4, t.evidence)
        ws.cell(row, 5, t.recommendation)
        ws.cell(row, 6, t.scope)
        ws.cell(row, 7, t.affected_entity)
        ws.cell(row, 8, t.timestamp)
        row += 1

    row += 1
    ws.cell(row, 1, "Rule Lint")
    row += 1
    for issue in lint_issues or ["No linter issues"]:
        ws.cell(row, 1, issue)
        row += 1


def export_rules(wb, path: Path):
    rules = table_rows(wb["Rules_Authoring"], "tblRules")
    payload = {
        "workbook": str(DEFAULT_WORKBOOK),
        "exported_at": dt.datetime.now().isoformat(),
        "rules": rules,
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    (LOG_DIR / "rules_export.log").write_text(f"{dt.datetime.now().isoformat()} exported {len(rules)} rules\n", encoding="utf-8")


def analyze(workbook_path: Path, rules_path: Path, export_only: bool = False):
    wb = load_workbook(workbook_path, keep_vba=True)
    if export_only:
        export_rules(wb, rules_path)
        wb.save(workbook_path)
        return

    schedule_rows = table_rows(wb["Schedule_Entry"], "tblSchedule")
    hourly_rows = table_rows(wb["Hourly_Log"], "tblHourly")
    downtime_rows = table_rows(wb["Downtime_Log"], "tblDowntime")
    standards_rows = table_rows(wb["Parameters"], "tblStandards")

    rules, source = select_rules(wb, rules_path)
    lint_issues = lint_rules(rules)
    triggers = evaluate_rules(rules, schedule_rows, hourly_rows, downtime_rows, standards_rows)

    missing_schedule = sum(1 for r in hourly_rows if missing_schedule_for_hourly(schedule_rows, r.get("Line"), parse_dt(r.get("HourEndingDT")) or dt.datetime.now()))
    missing_stds = sum(1 for r in hourly_rows if missing_standard(r.get("Line"), r.get("SKU_Resolved"), {(s.get("Line"), s.get("SKU")) for s in standards_rows}))

    sections = {
        "Data Quality": [
            f"Hourly rows: {len(hourly_rows)}",
            f"Downtime rows: {len(downtime_rows)}",
            f"Rules source: {source}",
        ],
        "Schedule Integrity": [f"Hourly rows without schedule: {missing_schedule}"],
        "Standards Coverage": [f"Rows missing standards: {missing_stds}"],
        "Operational Risks": [f"Triggered prompts: {len(triggers)}"],
        "Recommended Actions (ranked)": [f"{t.severity}: {t.recommendation} ({t.affected_entity})" for t in triggers[:10]],
    }
    write_analysis_report(wb, sections, triggers, lint_issues)
    wb.save(workbook_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--rules", default=str(DEFAULT_RULES_JSON))
    parser.add_argument("--export-rules", action="store_true")
    args = parser.parse_args()

    analyze(Path(args.workbook), Path(args.rules), export_only=args.export_rules)
    print("Analyze complete")


if __name__ == "__main__":
    main()
