#!/usr/bin/env python3
"""
consolidate_schedules.py

Reads a "Daily Production Schedule" workbook (one sheet per date) and produces
a new workbook organized by production line, with summary, data-issue log,
and README sheets.

Author: Claude (automated)
"""

import json
import re
import datetime
from collections import defaultdict
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ──────────────────────────────────────────────────────────────────────
# CONFIG
# ──────────────────────────────────────────────────────────────────────
INPUT_PATH  = "/mnt/data/Daily Production Schedule 8.8.25.xlsx"
OUTPUT_PATH = "/mnt/data/Production_Schedule_By_Line.xlsx"

KNOWN_HEADER_COLS = {
    # column index -> canonical field name (1-based)
    3: "Cases",
    4: "Shifts",
    5: "Cases Completed",
    6: "Target per Shift",
    7: "% Complete",
    8: "Notes",
    12: "Work order made",
}

LINE_NUMBERS = [1, 2, 3, 4, 5]

# ──────────────────────────────────────────────────────────────────────
# DATA STRUCTURES
# ──────────────────────────────────────────────────────────────────────
issues = []  # list of dicts for "Assumptions & Data Issues"

def log_issue(severity, sheet_name, date_val, line, row_ref, field, problem, action):
    issues.append({
        "Issue_ID": len(issues) + 1,
        "Severity": severity,
        "SheetName": sheet_name,
        "Date": str(date_val) if date_val else "",
        "Line": line,
        "RowRef": row_ref,
        "Field": field,
        "Problem": problem,
        "ActionTaken": action,
    })


# ──────────────────────────────────────────────────────────────────────
# DATE PARSING
# ──────────────────────────────────────────────────────────────────────
def parse_date_from_sheet_name(name):
    """Try to parse a date from a sheet name like '8.6.25', '08.13.2025', '10.01.25'."""
    name = name.strip()
    # Try MM.DD.YYYY
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', name)
    if m:
        month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    # Try M.D.YY
    m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{2})$', name)
    if m:
        month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year += 2000
        try:
            return datetime.date(year, month, day)
        except ValueError:
            return None
    return None


def extract_date_from_sheet(ws, sheet_name):
    """
    Look for a real datetime in the header rows (rows 1-5, columns A-C).
    Fall back to parsing the sheet name.
    """
    real_date = None
    for row_idx in range(1, min(ws.max_row + 1, 6)):
        for col_idx in range(1, 4):
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, datetime.datetime):
                real_date = val.date()
                break
        if real_date:
            break

    sheet_date = parse_date_from_sheet_name(sheet_name)

    if real_date:
        # Prefer real date; note if sheet name disagrees
        if sheet_date and real_date != sheet_date:
            log_issue("Warning", sheet_name, real_date, "", "",
                      "Date", f"Cell date {real_date} differs from sheet name date {sheet_date}",
                      f"Using cell date {real_date}")
        return real_date
    elif sheet_date:
        log_issue("Info", sheet_name, sheet_date, "", "",
                  "Date", "No cell date found; parsed from sheet name",
                  f"Using parsed date {sheet_date}")
        return sheet_date
    else:
        log_issue("Error", sheet_name, None, "", "",
                  "Date", "Cannot determine date from cell or sheet name",
                  "Skipping sheet")
        return None


# ──────────────────────────────────────────────────────────────────────
# DETECT EXTRA HEADER COLUMNS
# ──────────────────────────────────────────────────────────────────────
def detect_extra_columns(ws):
    """
    Scan the header row for column headers beyond the known set.
    Returns a dict {col_index: header_name} for any additional columns found.
    Also specifically look for "Work order made" column position.
    """
    extra_cols = {}
    work_order_col = 12  # default

    for row_idx in range(1, min(ws.max_row + 1, 6)):
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            val_str = str(val).strip().lower()
            if val_str == "work order made":
                work_order_col = col_idx
            # Check if this is a header row (has "Cases" in col C area)
            if col_idx in KNOWN_HEADER_COLS:
                continue
            if col_idx <= 2:
                continue
            # It's beyond our known columns — record it
            val_clean = str(ws.cell(row=row_idx, column=col_idx).value).strip()
            if val_clean and val_clean != '\xa0' and not isinstance(ws.cell(row=row_idx, column=col_idx).value, datetime.datetime):
                # Check it looks like a header (text, not a number)
                try:
                    float(val_clean)
                except (ValueError, TypeError):
                    extra_cols[col_idx] = val_clean

    return extra_cols, work_order_col


# ──────────────────────────────────────────────────────────────────────
# CELL VALUE HELPERS
# ──────────────────────────────────────────────────────────────────────
def clean_val(val):
    """Strip non-breaking spaces and whitespace; return None for empty."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.replace('\xa0', '').strip()
        return val if val else None
    return val


def to_numeric(val, field_name, sheet_name, date_val, line, row_ref):
    """Convert a value to float if possible; log warning if coercion fails."""
    val = clean_val(val)
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        # Remove commas
        val = val.replace(',', '').strip()
        try:
            return float(val)
        except (ValueError, TypeError):
            log_issue("Warning", sheet_name, date_val, line, row_ref,
                      field_name, f"Non-numeric value '{val}'",
                      "Left blank")
            return None
    return None


# ──────────────────────────────────────────────────────────────────────
# SKU PARSING
# ──────────────────────────────────────────────────────────────────────
def parse_sku(raw_text, sheet_name, date_val, line, row_ref):
    """
    Extract SKU code (first 6+ digit number) and description from raw text.
    Returns (sku, description, raw_text).
    """
    if raw_text is None:
        return None, None, None

    raw_str = str(raw_text).strip()
    if not raw_str:
        return None, None, raw_str

    # Find all sequences of 6+ digits
    candidates = re.findall(r'\b(\d{6,})\b', raw_str)

    if not candidates:
        # Try without word boundary (sometimes SKU is at start with no space)
        candidates = re.findall(r'(\d{6,})', raw_str)

    if not candidates:
        # Special case: short SKU like "1571" (4 digits)
        candidates_short = re.findall(r'\b(\d{4,})\b', raw_str)
        if candidates_short:
            sku = candidates_short[0]
            if len(candidates_short) > 1:
                log_issue("Warning", sheet_name, date_val, line, row_ref,
                          "SKU", f"Multiple short number candidates in '{raw_str}'; using first: {sku}",
                          f"Used SKU={sku}")
            log_issue("Info", sheet_name, date_val, line, row_ref,
                      "SKU", f"Short SKU ({len(sku)} digits) detected: {sku}",
                      f"Used SKU={sku}")
        else:
            log_issue("Warning", sheet_name, date_val, line, row_ref,
                      "SKU", f"No numeric SKU found in '{raw_str}'",
                      "Left SKU blank")
            return None, None, raw_str
    else:
        sku = candidates[0]
        if len(candidates) > 1:
            log_issue("Warning", sheet_name, date_val, line, row_ref,
                      "SKU", f"Multiple 6+ digit candidates in '{raw_str}'; using first: {sku}",
                      f"Used SKU={sku}")

    # Description extraction
    # Strategy:
    #   1. If " / " exists (explicit separator), description = everything after " / "
    #   2. Else if SKU is immediately followed by "/" (no space), description = after that "/"
    #   3. Else description = everything after the SKU token (including pack sizes like "1/8-15.25OZ ...")
    description = None
    sku_pos = raw_str.find(sku)
    after_sku = raw_str[sku_pos + len(sku):] if sku_pos >= 0 else ""

    if ' / ' in raw_str:
        sep_pos = raw_str.find(' / ')
        desc_candidate = raw_str[sep_pos + 3:].strip()
        if desc_candidate:
            description = desc_candidate
    elif after_sku.startswith('/'):
        # SKU immediately followed by "/" (e.g., "2001427/24/15.25OZ_DMWKGCRN")
        desc_candidate = after_sku[1:].strip()
        if desc_candidate:
            description = desc_candidate
    else:
        desc_candidate = after_sku.strip()
        if desc_candidate:
            description = desc_candidate

    return sku, description, raw_str


# ──────────────────────────────────────────────────────────────────────
# MAIN EXTRACTION
# ──────────────────────────────────────────────────────────────────────
def is_line_header(cell_a_val):
    """Check if column A value is a line header like 'Line 1', 'Line 2 ', etc."""
    if cell_a_val is None:
        return False
    s = str(cell_a_val).strip()
    m = re.match(r'^Line\s+(\d+)\s*$', s, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return False


def is_schedule_row_line_num(cell_a_val):
    """Check if column A has a numeric line number (1-5)."""
    val = clean_val(cell_a_val)
    if val is None:
        return None
    if isinstance(val, (int, float)):
        n = int(val)
        if 1 <= n <= 5:
            return n
    if isinstance(val, str):
        try:
            n = int(val.strip())
            if 1 <= n <= 5:
                return n
        except (ValueError, TypeError):
            pass
    return None


def extract_all_rows(wb):
    """
    Main extraction: iterate all sheets, find line sections, extract schedule rows.
    Returns dict: {line_num: [list of row dicts]}
    """
    all_rows = defaultdict(list)
    sheets_processed = 0
    sheets_skipped = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract date
        date_val = extract_date_from_sheet(ws, sheet_name)
        if date_val is None:
            sheets_skipped += 1
            continue

        sheets_processed += 1

        # Detect extra columns and work order column position
        extra_cols_map, work_order_col = detect_extra_columns(ws)

        # First pass: find all line header row positions
        line_header_rows = []
        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            line_num = is_line_header(cell_a)
            if line_num:
                line_header_rows.append((row_idx, line_num))

        # Determine current_line context for each row
        # Build a mapping: row_idx -> current_line_context
        # (the most recent line header above this row)
        current_line_by_row = {}
        current_line = None
        header_row_set = {r for r, _ in line_header_rows}

        for row_idx in range(1, ws.max_row + 1):
            if row_idx in header_row_set:
                for hr, ln in line_header_rows:
                    if hr == row_idx:
                        current_line = ln
                        break
            current_line_by_row[row_idx] = current_line

        # Second pass: extract schedule rows
        for row_idx in range(1, ws.max_row + 1):
            if row_idx in header_row_set:
                continue  # skip header rows

            cell_a_val = ws.cell(row=row_idx, column=1).value
            line_num = is_schedule_row_line_num(cell_a_val)

            if line_num is None:
                # Not a schedule row — but check if it has Target per Shift on a filler row
                target_val = clean_val(ws.cell(row=row_idx, column=6).value)
                col_b_val = clean_val(ws.cell(row=row_idx, column=2).value)
                col_c_val = clean_val(ws.cell(row=row_idx, column=3).value)
                col_d_val = clean_val(ws.cell(row=row_idx, column=4).value)
                col_e_val = clean_val(ws.cell(row=row_idx, column=5).value)

                if target_val is not None and col_b_val is None and col_c_val is None and col_d_val is None:
                    log_issue("Info", sheet_name, date_val,
                              current_line_by_row.get(row_idx, ""),
                              row_idx, "Target_Per_Shift",
                              f"Target per Shift ({target_val}) on filler row",
                              "Ignored filler row")
                continue

            # We have a schedule row with line_num
            col_b_raw = ws.cell(row=row_idx, column=2).value
            col_c = ws.cell(row=row_idx, column=3).value
            col_d = ws.cell(row=row_idx, column=4).value
            col_e = ws.cell(row=row_idx, column=5).value
            col_f = ws.cell(row=row_idx, column=6).value
            col_g = ws.cell(row=row_idx, column=7).value
            col_h = ws.cell(row=row_idx, column=8).value

            # Check if col_b is empty and other fields too
            col_b_clean = clean_val(col_b_raw)
            col_c_clean = clean_val(col_c)
            col_d_clean = clean_val(col_d)

            if col_b_clean is None and col_c_clean is None and col_d_clean is None:
                # Filler row even though it has line number in A
                target_check = clean_val(col_f)
                if target_check is not None:
                    log_issue("Info", sheet_name, date_val, line_num, row_idx,
                              "Target_Per_Shift",
                              f"Target ({target_check}) on row with line num but no SKU/Cases/Shifts",
                              "Ignored filler row")
                continue

            if col_b_clean is None:
                # Has Cases or Shifts but no SKU text
                log_issue("Warning", sheet_name, date_val, line_num, row_idx,
                          "SKU", "Missing SKU text but other fields present",
                          "Extracted with blank SKU")

            # Parse SKU
            sku, description, raw_text = parse_sku(col_b_clean, sheet_name, date_val, line_num, row_idx)

            # Parse numeric fields
            cases_planned = to_numeric(col_c, "Cases_Planned", sheet_name, date_val, line_num, row_idx)
            shifts_planned = to_numeric(col_d, "Shifts_Planned", sheet_name, date_val, line_num, row_idx)
            cases_completed = to_numeric(col_e, "Cases_Completed", sheet_name, date_val, line_num, row_idx)
            target_per_shift = to_numeric(col_f, "Target_Per_Shift", sheet_name, date_val, line_num, row_idx)

            # Work order
            work_order_val = clean_val(ws.cell(row=row_idx, column=work_order_col).value)

            # Notes
            notes_val = clean_val(col_h)

            # Percent complete — compute, don't copy
            pct_complete = None
            if cases_planned is not None and cases_planned > 0 and cases_completed is not None:
                pct_complete = cases_completed / cases_planned

            # Extra fields: scan all columns beyond known ones
            extra_fields = {}
            for col_idx in range(1, ws.max_column + 1):
                if col_idx in (1, 2, 3, 4, 5, 6, 7, 8, work_order_col):
                    continue
                cell_val = clean_val(ws.cell(row=row_idx, column=col_idx).value)
                if cell_val is not None:
                    header = extra_cols_map.get(col_idx, get_column_letter(col_idx))
                    extra_fields[header] = cell_val

            extra_json = json.dumps(extra_fields) if extra_fields else ""

            # Sanity checks
            if cases_planned is not None and cases_planned < 0:
                log_issue("Warning", sheet_name, date_val, line_num, row_idx,
                          "Cases_Planned", f"Negative value: {cases_planned}",
                          "Kept as-is")
            if shifts_planned is not None and shifts_planned < 0:
                log_issue("Warning", sheet_name, date_val, line_num, row_idx,
                          "Shifts_Planned", f"Negative value: {shifts_planned}",
                          "Kept as-is")
            if shifts_planned is not None and shifts_planned > 0 and target_per_shift is None:
                log_issue("Warning", sheet_name, date_val, line_num, row_idx,
                          "Target_Per_Shift", "Missing when Shifts > 0",
                          "Left blank")

            row_data = {
                "Date": date_val,
                "SourceSheet": sheet_name,
                "Line": line_num,
                "SKU": sku,
                "SKU_RawText": raw_text,
                "Description": description,
                "Cases_Planned": cases_planned,
                "Shifts_Planned": shifts_planned,
                "Target_Per_Shift": target_per_shift,
                "Cases_Completed": cases_completed,
                "Percent_Complete": pct_complete,
                "Notes": notes_val,
                "WorkOrderMade": work_order_val,
                "ExtraFields_JSON": extra_json,
            }

            all_rows[line_num].append(row_data)

    return all_rows, sheets_processed, sheets_skipped


# ──────────────────────────────────────────────────────────────────────
# DUPLICATE DETECTION
# ──────────────────────────────────────────────────────────────────────
def detect_duplicates(all_rows):
    """Find duplicate rows by (Date, Line, SKU, Cases_Planned, Shifts_Planned)."""
    seen = defaultdict(list)
    for line_num, rows in all_rows.items():
        for i, r in enumerate(rows):
            key = (str(r["Date"]), r["Line"], r["SKU"], r["Cases_Planned"], r["Shifts_Planned"])
            seen[key].append((line_num, i, r))

    dup_count = 0
    for key, entries in seen.items():
        if len(entries) > 1:
            for line_num, idx, r in entries:
                log_issue("Warning", r["SourceSheet"], r["Date"], r["Line"],
                          f"data row {idx+1}", "Duplicate",
                          f"Duplicate key: Date={r['Date']}, Line={r['Line']}, SKU={r['SKU']}, "
                          f"Cases={r['Cases_Planned']}, Shifts={r['Shifts_Planned']}",
                          "Kept (not deleted)")
                dup_count += 1
    return dup_count


# ──────────────────────────────────────────────────────────────────────
# OUTPUT WORKBOOK
# ──────────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

LINE_SHEET_COLUMNS = [
    "Date", "SourceSheet", "Line", "SKU", "SKU_RawText", "Description",
    "Cases_Planned", "Shifts_Planned", "Target_Per_Shift", "Cases_Completed",
    "Percent_Complete", "Notes", "WorkOrderMade", "ExtraFields_JSON",
]


def write_readme(wb_out, all_rows, sheets_processed, sheets_skipped):
    ws = wb_out.create_sheet("README", 0)

    total_rows = sum(len(rows) for rows in all_rows.values())
    per_line = {ln: len(all_rows.get(ln, [])) for ln in LINE_NUMBERS}

    readme_lines = [
        ["Production Schedule Consolidation — README"],
        [],
        ["What this workbook contains:"],
        ["This workbook was generated by an automated script that reads a Daily Production"],
        ["Schedule workbook (one sheet per date) and reorganizes it by production line."],
        [],
        ["How dates were detected:"],
        ["1. First, the script looks for a real datetime cell in rows 1-5, columns A-C of each sheet."],
        ["2. If not found, it parses the sheet name (formats: M.D.YY, MM.DD.YYYY)."],
        ["3. If neither works, the sheet is skipped and logged as an Error."],
        [],
        ["How lines were detected:"],
        ["A 'Line N' header in column A starts a section. Data rows have the line number (1-5) in column A."],
        ["If data rows appear before any Line header, they are still captured using the line number in col A."],
        [],
        ["How SKUs were detected:"],
        ["The first 6+ digit number in the column B text is extracted as the SKU code."],
        ["If only 4-5 digit numbers exist, the first is used (with a logged Info)."],
        [],
        ["Known limitations:"],
        ["- Description parsing is best-effort; ambiguous cases are left blank."],
        ["- Some sheets' cell dates differ from their sheet names (warnings logged)."],
        ["- Filler rows with only Target per Shift are ignored (logged as Info)."],
        ["- '#DIV/0!' in % Complete is ignored; Percent_Complete is recomputed."],
        [],
        ["Processing statistics:"],
        [f"  Sheets processed: {sheets_processed}"],
        [f"  Sheets skipped: {sheets_skipped}"],
        [f"  Total schedule rows extracted: {total_rows}"],
    ]

    for ln in LINE_NUMBERS:
        readme_lines.append([f"    Line {ln}: {per_line[ln]} rows"])

    readme_lines += [
        [],
        [f"  Issues logged: {len(issues)}"],
        [f"    Info: {sum(1 for i in issues if i['Severity'] == 'Info')}"],
        [f"    Warning: {sum(1 for i in issues if i['Severity'] == 'Warning')}"],
        [f"    Error: {sum(1 for i in issues if i['Severity'] == 'Error')}"],
    ]

    for i, line_data in enumerate(readme_lines, 1):
        for j, val in enumerate(line_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            if i == 1:
                cell.font = Font(bold=True, size=14)

    ws.column_dimensions['A'].width = 90
    ws.freeze_panes = "A2"


def write_issues(wb_out):
    ws = wb_out.create_sheet("Assumptions & Data Issues")

    headers = ["Issue_ID", "Severity", "SheetName", "Date", "Line",
               "RowRef", "Field", "Problem", "ActionTaken"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, issue in enumerate(issues, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=issue.get(key, ""))

    # Auto-fit (approximate)
    col_widths = {"Issue_ID": 10, "Severity": 10, "SheetName": 16, "Date": 12,
                  "Line": 6, "RowRef": 8, "Field": 18, "Problem": 60, "ActionTaken": 40}
    for col_idx, key in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(issues) + 1)}"


def write_summary(wb_out, all_rows):
    ws = wb_out.create_sheet("Summary")

    headers = ["Date", "Line", "Total_Planned_Cases", "Total_Planned_Shifts",
               "Total_Completed_Cases", "Avg_Pct_Complete_Weighted", "Count_SKUs"]

    # Build summary data
    summary_data = defaultdict(lambda: {
        "cases_planned": 0, "shifts_planned": 0, "completed": 0,
        "weighted_pct_num": 0, "weighted_pct_den": 0, "count": 0
    })

    for line_num, rows in all_rows.items():
        for r in rows:
            key = (r["Date"], r["Line"])
            s = summary_data[key]
            cp = r["Cases_Planned"] or 0
            sp = r["Shifts_Planned"] or 0
            cc = r["Cases_Completed"] or 0
            s["cases_planned"] += cp
            s["shifts_planned"] += sp
            s["completed"] += cc
            s["count"] += 1
            if r["Percent_Complete"] is not None and cp > 0:
                s["weighted_pct_num"] += r["Percent_Complete"] * cp
                s["weighted_pct_den"] += cp

    # Sort by date, then line
    sorted_keys = sorted(summary_data.keys(), key=lambda k: (k[0], k[1]))

    # Overall totals
    overall = {"cases_planned": 0, "shifts_planned": 0, "completed": 0,
               "weighted_pct_num": 0, "weighted_pct_den": 0, "count": 0}
    for s in summary_data.values():
        overall["cases_planned"] += s["cases_planned"]
        overall["shifts_planned"] += s["shifts_planned"]
        overall["completed"] += s["completed"]
        overall["weighted_pct_num"] += s["weighted_pct_num"]
        overall["weighted_pct_den"] += s["weighted_pct_den"]
        overall["count"] += s["count"]

    # Write overall totals at top
    ws.cell(row=1, column=1, value="OVERALL TOTALS").font = Font(bold=True, size=12)
    overall_headers = ["", "", "Total_Planned_Cases", "Total_Planned_Shifts",
                       "Total_Completed_Cases", "Avg_Pct_Complete_Weighted", "Count_SKUs"]
    for col_idx, h in enumerate(overall_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = HEADER_FONT

    overall_avg_pct = (overall["weighted_pct_num"] / overall["weighted_pct_den"]
                       if overall["weighted_pct_den"] > 0 else None)
    overall_vals = ["", "", overall["cases_planned"], overall["shifts_planned"],
                    overall["completed"], overall_avg_pct, overall["count"]]
    for col_idx, v in enumerate(overall_vals, 1):
        cell = ws.cell(row=3, column=col_idx, value=v)
        if col_idx == 6 and v is not None:
            cell.number_format = '0.0%'

    # Blank row
    start_row = 5

    # Column headers for detail
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, key in enumerate(sorted_keys, start_row + 1):
        date_val, line_num = key
        s = summary_data[key]
        avg_pct = (s["weighted_pct_num"] / s["weighted_pct_den"]
                   if s["weighted_pct_den"] > 0 else None)

        ws.cell(row=i, column=1, value=datetime.datetime(date_val.year, date_val.month, date_val.day)).number_format = 'm/d/yyyy'
        ws.cell(row=i, column=2, value=line_num)
        ws.cell(row=i, column=3, value=s["cases_planned"])
        ws.cell(row=i, column=4, value=s["shifts_planned"])
        ws.cell(row=i, column=5, value=s["completed"])
        cell_pct = ws.cell(row=i, column=6, value=avg_pct)
        if avg_pct is not None:
            cell_pct.number_format = '0.0%'
        ws.cell(row=i, column=7, value=s["count"])

    # Widths
    col_widths_sum = [12, 6, 20, 20, 22, 26, 12]
    for col_idx, w in enumerate(col_widths_sum, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = f"A{start_row + 1}"

    # Reconciliation: verify totals match line sheets
    line_totals = {"cases_planned": 0, "shifts_planned": 0, "completed": 0, "count": 0}
    for line_num in LINE_NUMBERS:
        for r in all_rows.get(line_num, []):
            line_totals["cases_planned"] += (r["Cases_Planned"] or 0)
            line_totals["shifts_planned"] += (r["Shifts_Planned"] or 0)
            line_totals["completed"] += (r["Cases_Completed"] or 0)
            line_totals["count"] += 1

    if line_totals["cases_planned"] != overall["cases_planned"]:
        log_issue("Error", "Summary", "", "", "", "Reconciliation",
                  f"Summary cases_planned ({overall['cases_planned']}) != line sheets sum ({line_totals['cases_planned']})",
                  "Check for data issues")
    if line_totals["count"] != overall["count"]:
        log_issue("Error", "Summary", "", "", "", "Reconciliation",
                  f"Summary row count ({overall['count']}) != line sheets sum ({line_totals['count']})",
                  "Check for data issues")

    return overall


def write_line_sheet(wb_out, line_num, rows):
    sheet_name = f"Line {line_num}"
    ws = wb_out.create_sheet(sheet_name)

    # Header row
    for col_idx, h in enumerate(LINE_SHEET_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Sort rows by Date, then by original order (stable sort preserves extraction order)
    rows_sorted = sorted(rows, key=lambda r: r["Date"])

    for row_idx, r in enumerate(rows_sorted, 2):
        # Date
        d = r["Date"]
        cell_date = ws.cell(row=row_idx, column=1,
                            value=datetime.datetime(d.year, d.month, d.day))
        cell_date.number_format = 'm/d/yyyy'

        ws.cell(row=row_idx, column=2, value=r["SourceSheet"])
        ws.cell(row=row_idx, column=3, value=r["Line"])
        ws.cell(row=row_idx, column=4, value=r["SKU"])
        ws.cell(row=row_idx, column=5, value=r["SKU_RawText"])
        ws.cell(row=row_idx, column=6, value=r["Description"])

        # Cases_Planned — integer format
        cp = r["Cases_Planned"]
        cell_cp = ws.cell(row=row_idx, column=7, value=cp)
        if cp is not None:
            cell_cp.number_format = '#,##0'

        sp = r["Shifts_Planned"]
        cell_sp = ws.cell(row=row_idx, column=8, value=sp)
        if sp is not None:
            cell_sp.number_format = '#,##0'

        tps = r["Target_Per_Shift"]
        cell_tps = ws.cell(row=row_idx, column=9, value=tps)
        if tps is not None:
            cell_tps.number_format = '#,##0'

        cc = r["Cases_Completed"]
        cell_cc = ws.cell(row=row_idx, column=10, value=cc)
        if cc is not None:
            cell_cc.number_format = '#,##0'

        pct = r["Percent_Complete"]
        cell_pct = ws.cell(row=row_idx, column=11, value=pct)
        if pct is not None:
            cell_pct.number_format = '0.0%'

        ws.cell(row=row_idx, column=12, value=r["Notes"])
        ws.cell(row=row_idx, column=13, value=r["WorkOrderMade"])
        ws.cell(row=row_idx, column=14, value=r["ExtraFields_JSON"])

    # Column widths
    col_widths_line = {
        1: 12, 2: 18, 3: 6, 4: 10, 5: 50, 6: 40,
        7: 14, 8: 14, 9: 16, 10: 16, 11: 16, 12: 40, 13: 16, 14: 40,
    }
    for col_idx, w in col_widths_line.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = min(w, 50)

    # Freeze panes: top row + first 3 columns
    ws.freeze_panes = "D2"

    # Create Excel Table
    last_row = max(2, len(rows_sorted) + 1)
    last_col_letter = get_column_letter(len(LINE_SHEET_COLUMNS))
    table_ref = f"A1:{last_col_letter}{last_row}"
    table_name = f"tblLine{line_num}Schedule"

    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Conditional formatting
    # Range for data rows (excluding header)
    pct_range = f"K2:K{last_row}"
    tps_range = f"I2:I{last_row}"
    sku_range = f"D2:D{last_row}"

    # Percent_Complete < 80% highlight (yellow)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.conditional_formatting.add(pct_range,
        CellIsRule(operator='lessThan', formula=['0.8'], fill=yellow_fill))

    # Missing Target_Per_Shift highlight (light red)
    light_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(tps_range,
        CellIsRule(operator='equal', formula=['""'], fill=light_red))

    # Missing SKU highlight (light orange)
    light_orange = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
    ws.conditional_formatting.add(sku_range,
        CellIsRule(operator='equal', formula=['""'], fill=light_orange))

    # Also update the sorted rows back (for consistency)
    all_rows_for_line = rows_sorted
    return len(rows_sorted)


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────
def main():
    print("Loading input workbook...")
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)

    print("Extracting schedule rows...")
    all_rows, sheets_processed, sheets_skipped = extract_all_rows(wb)

    print("Detecting duplicates...")
    dup_count = detect_duplicates(all_rows)

    print("Creating output workbook...")
    wb_out = openpyxl.Workbook()
    # Remove default sheet
    wb_out.remove(wb_out.active)

    # Write sheets
    write_readme(wb_out, all_rows, sheets_processed, sheets_skipped)
    write_issues(wb_out)
    write_summary(wb_out, all_rows)

    for ln in LINE_NUMBERS:
        rows = all_rows.get(ln, [])
        row_count = write_line_sheet(wb_out, ln, rows)
        print(f"  Line {ln}: {row_count} rows written")

    # Save
    print(f"Saving to {OUTPUT_PATH}...")
    wb_out.save(OUTPUT_PATH)
    print("Done!")

    # Print run log
    print("\n" + "=" * 60)
    print("RUN LOG")
    print("=" * 60)
    print(f"Sheets processed: {sheets_processed}")
    print(f"Sheets skipped:   {sheets_skipped}")

    total = 0
    for ln in LINE_NUMBERS:
        cnt = len(all_rows.get(ln, []))
        total += cnt
        print(f"  Line {ln}: {cnt} rows")
    print(f"  TOTAL: {total} rows")
    print(f"Duplicates found: {dup_count}")

    sev_counts = defaultdict(int)
    for i in issues:
        sev_counts[i["Severity"]] += 1
    print(f"Issues logged: {len(issues)}")
    for sev in ["Info", "Warning", "Error"]:
        print(f"  {sev}: {sev_counts[sev]}")


if __name__ == "__main__":
    main()
