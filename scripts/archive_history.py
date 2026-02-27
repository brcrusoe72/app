#!/usr/bin/env python3
"""Archive workbook logs into SQLite with RowID dedupe."""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = REPO_ROOT / "excel" / "Shift_Flight_Deck.xlsm"
DB_PATH = REPO_ROOT / "data" / "history.sqlite"


def table_rows(ws, table_name):
    tab = ws.tables[table_name]
    min_cell, max_cell = tab.ref.split(":")
    min_col = ord(min_cell[0]) - 64
    min_row = int(min_cell[1:])
    max_col = ord(max_cell[0]) - 64
    max_row = int(max_cell[1:])
    headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
    rows = []
    for r in range(min_row + 1, max_row + 1):
        vals = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]
        if any(v not in (None, "") for v in vals):
            rows.append(dict(zip(headers, vals)))
    return rows


def ensure_tables(conn: sqlite3.Connection):
    conn.execute("CREATE TABLE IF NOT EXISTS schedule_log (RowID TEXT PRIMARY KEY, payload TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS hourly_log (RowID TEXT PRIMARY KEY, payload TEXT)")
    conn.execute("CREATE TABLE IF NOT EXISTS downtime_log (RowID TEXT PRIMARY KEY, payload TEXT)")


def upsert_rows(conn, table_name, rows):
    for row in rows:
        conn.execute(f"INSERT OR REPLACE INTO {table_name}(RowID, payload) VALUES (?, ?)", (row.get("RowID"), str(row)))


def archive(workbook_path: Path, clear_current: bool):
    wb = load_workbook(workbook_path, keep_vba=True)
    schedule = table_rows(wb["Schedule_Entry"], "tblSchedule")
    hourly = table_rows(wb["Hourly_Log"], "tblHourly")
    downtime = table_rows(wb["Downtime_Log"], "tblDowntime")

    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    ensure_tables(conn)
    upsert_rows(conn, "schedule_log", schedule)
    upsert_rows(conn, "hourly_log", hourly)
    upsert_rows(conn, "downtime_log", downtime)
    conn.commit()
    conn.close()

    if clear_current:
        for ws_name, table in [("Schedule_Entry", "tblSchedule"), ("Hourly_Log", "tblHourly"), ("Downtime_Log", "tblDowntime")]:
            ws = wb[ws_name]
            tab = ws.tables[table]
            _, max_cell = tab.ref.split(":")
            max_row = int(max_cell[1:])
            if max_row > 1:
                ws.delete_rows(2, max_row - 1)
        wb.save(workbook_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--clear-current", action="store_true")
    args = parser.parse_args()
    archive(Path(args.workbook), args.clear_current)
    print("Archive complete")


if __name__ == "__main__":
    main()
